# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
from prettytable import PrettyTable

# Load in the workbook
wb = load_workbook('plan.xlsx')

sheet = wb['1 курс ИТ']
c = (sheet['O4'].value)
td = []

for i in range(5, 41):
    th = []
    for j in [2, 15, 16, 17, 18]:
        data = sheet.cell(row=i, column=j).value
        th.append(data)
    table = PrettyTable(th)
    print(table)