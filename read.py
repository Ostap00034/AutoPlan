# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
from prettytable import PrettyTable

# Load in the workbook
wb = load_workbook('plan.xlsx')

sheet = wb['1 курс ИТ']
c = (sheet['O4'].value)

th = []
td = []

for i in range(5, 41):
    temp_data = []
    for j in [2, 15, 16, 17, 18]:
        data = sheet.cell(row=i, column=j).value
        if i == 5:
            th.append(data)
            table = PrettyTable(th)
        else:
            if data == 'None':
                data = ' '
            temp_data.append(data)
    if i != 5:
        table.add_row(temp_data)
        
# print(sheet.cell(row=4, column=15).value)
print(table)