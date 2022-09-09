import requests
from bs4 import BeautifulSoup as BS
from xls2xlsx import XLS2XLSX
from openpyxl import load_workbook
from prettytable import PrettyTable
from datetime import datetime, date, time
import os

# Функции скачивания файла
def download_file_from_google_drive(id, destination):
    URL = "https://docs.google.com/uc?export=download"

    session = requests.Session()

    response = session.get(URL, params = { 'id' : id }, stream = True)
    token = get_confirm_token(response)

    if token:
        params = { 'id' : id, 'confirm' : token }
        response = session.get(URL, params = params, stream = True)

    save_response_content(response, destination)    

def get_confirm_token(response):
    for key, value in response.cookies.items():
        if key.startswith('download_warning'):
            return value

    return None

def save_response_content(response, destination):
    CHUNK_SIZE = 32768

    with open(destination, "wb") as f:
        for chunk in response.iter_content(CHUNK_SIZE):
            if chunk: # filter out keep-alive new chunks
                f.write(chunk)

# Ссылка сайта с ссылками на расписания
url = 'https://www.s-vfu.ru/universitet/rukovodstvo-i-struktura/instituty/imi/Vremennoe_raspisanie_IMI_na_1_2020-2021/'

# Получение HTML кода страницы
r = requests.get(url)
html = BS(r.content, 'html.parser')

# Поиск и нахождение ссылки на файлы расписания
for i in (html.select('a')):
    # Определение строки (соответсвующая ли?)
    s = str(i)
    if 'ИМИ Расписание учебных занятий на 1 полугодие ' in s:
        url_file = s.split('"')[1]
        if ('spreadsheets' in url_file):
            file_id = url_file.split('/')[5]
            # Скачивание файла
            if __name__ == "__main__":
                file_id = file_id
                destination = '1.xls'
                download_file_from_google_drive(file_id, destination)
# Обработка файла
# Конвертирование XLS в XLSX
x2x = XLS2XLSX("1.xls")
wb = x2x.to_xlsx("plan.xlsx")

# Открытие файла для чтения
wb = load_workbook('plan.xlsx')
# Выбор страницы

th = []
td = []

sheet = wb['1 курс_МО']
c = (sheet['O4'].value)

day = (datetime.weekday(datetime.now())) * 6
th = ['День недели', 'Время', 'Дисциплина', 'Преподователь', 'Тип', 'Кабинет']
for i in range(day + 5, day + 11):
    temp_data = []
    for j in [1, 2, 11, 12, 13, 14]:
        data = sheet.cell(row=i, column=j).value
        if data == None:
            data = ' '
        if i == day + 5:
            # th.append(data)
            table = PrettyTable(th)
        else:
            temp_data.append(data)
    if i != day + 5:
        table.add_row(temp_data)
        
# print(sheet.cell(row=4, column=15).value)
print(table)
os.remove('1.xls')
os.remove('plan.xlsx')
print(input())

